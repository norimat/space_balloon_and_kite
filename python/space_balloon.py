#!/usr/bin/env python
try:
    from picamera2 import Picamera2
    from picamera2.encoders import H264Encoder
    import smbus2
    import bme280
    import qwiic_icm20948
    import qwiic_i2c
    import serial
    import pynmea2
    import psutil
except ImportError:
    print("[Warn] The libraries required for reading sensor data from GPIO or related interfaces have not been imported.")
try:
    import folium
    from folium.plugins import TimestampedGeoJson
    import simplekml
    import pandas
    import matplotlib
    import numpy
    import cv2
    import openpyxl
    from openpyxl.styles import PatternFill , Alignment , Font , Border , Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("[Warn] The libraries required for analyzing sensor data have not been imported.")

import math
import subprocess
import time
import csv
import signal
import sys
import argparse
import threading
import multiprocessing
import shutil
import os
import re
import datetime
import glob
import struct
import json
import queue
import gc

########################################################################
class SensorWrapper:

    # bme280_cond             = threading.Condition()
    mpu6050_cond            = threading.Condition()
    icm20948_cond           = threading.Condition()
    camera_module_cond      = threading.Condition()
    powermonitor_cond       = threading.Condition()
    running                 = threading.Event()
    # bme280_ready            = False
    mpu6050_ready           = False
    icm20948_ready          = False
    camera_module_ready     = False
    powermonitor_ready      = False

    # init
    start_time                     = 0
    bme280_start_time              = 0
    bme280_end_time                = 0
    bme280_byte_0                  = 0
    bme280_byte_1                  = 0
    bme280_byte_2                  = 0
    bme280_byte_3                  = 0
    bme280_byte_4                  = 0
    bme280_byte_5                  = 0
    bme280_byte_6                  = 0
    bme280_byte_7                  = 0
    mpu6050_start_time             = 0
    mpu6050_end_time               = 0
    mpu6050_byte_0                 = 0
    mpu6050_byte1                  = 0
    mpu6050_byte2                  = 0
    mpu6050_byte3                  = 0
    mpu6050_byte4                  = 0
    mpu6050_byte5                  = 0
    mpu6050_byte6                  = 0
    mpu6050_byte7                  = 0
    mpu6050_byte8                  = 0
    mpu6050_byte9                  = 0
    mpu6050_byte10                 = 0
    mpu6050_byte11                 = 0
    mpu6050_byte12                 = 0
    mpu6050_byte13                 = 0
    icm20948_start_time            = 0
    icm20948_end_time              = 0
    icm20948_axRaw                 = 0
    icm20948_ayRaw                 = 0
    icm20948_azRaw                 = 0
    icm20948_gxRaw                 = 0
    icm20948_gyRaw                 = 0
    icm20948_gzRaw                 = 0
    icm20948_mxRaw                 = 0
    icm20948_myRaw                 = 0
    icm20948_mzRaw                 = 0
    icm20948_tmpRaw                = 0
    ivk172_latitude                = 0
    ivk172_longitude               = 0
    ivk172_altitude                = 0
    ivk172_altitude_units          = 0
    ivk172_num_sats                = 0
    ivk172_datestamp               = 0
    ivk172_timestamp               = 0
    ivk172_spd_over_grnd           = 0
    ivk172_true_course             = 0
    ivk172_true_track              = 0
    ivk172_spd_over_grnd_kmph      = 0
    ivk172_pdop                    = 0
    ivk172_hdop                    = 0
    ivk172_vdo                     = 0
    ivk172_num_sv_in_view          = 0
    ivk172_frame                   = 0
    powermonitor_start_time        = 0
    powermonitor_end_time          = 0
    powermonitor_voltage           = 0
    powermonitor_throttled         = 0
    powermonitor_cpu               = 0
    powermonitor_mem_used_B        = 0
    powermonitor_mem_total_B       = 0
    powermonitor_mem_available_B   = 0
    powermonitor_mem_percent_used  = 0
    powermonitor_temp              = 0
    powermonitor_disk_used_B       = 0
    powermonitor_disk_total_B      = 0
    powermonitor_disk_free_B       = 0
    powermonitor_disk_percent_used = 0

    def __init__( self , argv ):
        print("[Info] Create an instance of the SensorWrapper class.")
        print("[Info] Current nice value is " + str(os.nice(0)) + ".")
        try:
            os.nice(-20)
            print("[Info] Priority set to maximum.")
            print("[Info] Current nice value is " + str(os.nice(0)) + ".")
        except PermissionError:
            print("[Warn] To run with higher priority, you must execute the process with root privileges.")
        self.argv                    = argv
        self.__bme280_bus            = None
        self.__mpu6050_bus           = None
        self.__camera_fa             = None
        self.__bme280_fa             = None
        self.__mpu6050_fa            = None
        self.__icm20948_fa           = None
        self.__powermonitor_fa       = None
        self.__gps_fa                = None
        self.__mode                  = None
        self.__json_output_dir       = None
        self.__csv_output_dir        = None
        self.__movie_output_dir      = None
        self.__icm20948_i2cbus       = None
        self.__bme280_i2cbus         = None
        self.__mpu6050_i2cbus        = None
        self.__gps_en                = None
        self.__powermonitor_en       = None
        self.__bme280_en             = None
        self.__mpu6050_en            = None
        self.__icm20948_en           = None
        self.__framerate             = None
        self.__framebuffer           = None
        self.__bitrate               = None
        self.__width                 = None
        self.__height                = None
        self.__csvbuffer             = None
        self.__gps_port              = None
        self.__bme280_addr           = None
        self.__mpu6050_addr          = None
        self.__icm20948_addr         = None
        self.__gps_csv               = None
        self.__bme280_csv            = None
        self.__mpu6050_csv           = None
        self.__icm20948_csv          = None
        self.__gps_interval          = None
        self.__bme280_interval       = None
        self.__analyzerDic           = {}

    def __handler( self , signum , frame ):
        SensorWrapper.running.clear()
        self.__picamera2.stop_encoder()
        self.__picamera2.stop()
        self.__camera_fa       .close()
        self.__bme280_en       and self.__bme280_fa       .close()
        self.__mpu6050_en      and self.__mpu6050_fa      .close()
        self.__icm20948_en     and self.__icm20948_fa     .close()
        self.__powermonitor_en and self.__powermonitor_fa .close()
        self.__gps_en          and self.__gps_fa          .close()
        self.__bme280_bus      .close()
        self.__mpu6050_bus     .close()
        SensorWrapper.camera_module_ready = True
        SensorWrapper.bme280_ready        = True
        SensorWrapper.mpu6050_ready       = True
        SensorWrapper.icm20948_ready      = True
        SensorWrapper.powermonitor_ready  = True
        SensorWrapper.bme280_cond              .notify()
        SensorWrapper.runningmpu6050_cond      .notify()
        SensorWrapper.runningicm20948_cond     .notify()
        SensorWrapper.runningcamera_module_cond.notify()
        SensorWrapper.runningpowermonitor_cond .notify()
        shutil.rmtree( './tmp' , ignore_errors=True )

    def __generate_empty_csvFile( self , csvFileName , data ):
        print("[Info] Create the  " + csvFileName + ".")
        fopen  = open( csvFileName , 'w' , newline='' , encoding='utf-8' )
        writer = csv.writer( fopen )
        writer.writerows( data )
        fopen.close()

    def __get_csvFile( self , csvFileName , csvbuffer ):
        fappend = open( csvFileName , 'a' , newline='' , encoding='utf-8' , buffering=csvbuffer )
        return fappend

    def __read_args( self ):
        parser = argparse.ArgumentParser( description='option' , formatter_class=argparse.RawTextHelpFormatter )
        parser.add_argument( '--mode'       , '-m' , default=0     , required=True       , help="" )
        #############################################################################################
        # Sensor Acquisition Mode Options
        parser.add_argument( '--json_output_dir'       , default="./"                        , help="" )
        parser.add_argument( '--csv_output_dir'        , default="./"                        , help="" )
        parser.add_argument( '--movie_output_dir'      , default="./"                        , help="" )
        parser.add_argument( '--gps'                   , default=False , action='store_true' , help="" )
        parser.add_argument( '--bme280'                , default=False , action='store_true' , help="" )
        parser.add_argument( '--mpu6050'               , default=False , action='store_true' , help="" )
        parser.add_argument( '--icm20948'              , default=False , action='store_true' , help="" )
        parser.add_argument( '--powermonitor'          , default=False , action='store_true' , help="" )
        parser.add_argument( '--icm20948_i2cbus'       , default=1                           , help="" )
        parser.add_argument( '--bme280_i2cbus'         , default=1                           , help="" )
        parser.add_argument( '--mpu6050_i2cbus'        , default=1                           , help="" )
        parser.add_argument( '--framerate'             , default="30"                        , help="" )
        parser.add_argument( '--framebuffer'           , default="4"                         , help="" )
        parser.add_argument( '--bitrate'               , default="8000000"                   , help="" )
        parser.add_argument( '--width'                 , default="1920"                      , help="" )
        parser.add_argument( '--height'                , default="1080"                      , help="" )
        parser.add_argument( '--csvbuffer'             , default="512"                       , help="" )
        parser.add_argument( '--gps_port'              , default="/dev/ttyACM0"              , help="" )
        parser.add_argument( '--gps_interval'          , default="5.0"                       , help="" )
        parser.add_argument( '--bme280_interval'       , default="5.0"                       , help="" )
        parser.add_argument( '--bme280_addr'           , default="0x76"                      , help="" )
        parser.add_argument( '--mpu6050_addr'          , default="0x68"                      , help="" )
        parser.add_argument( '--icm20948_addr'         , default="0x68"                      , help="" )
        #############################################################################################
        # Sensor Data Analysis Mode Options
        parser.add_argument( '--input_dir'      , '-i' , default="./"                        , help="" )
        parser.add_argument( '--calib_json'            , default="./mag_calib.json"          , help="" )
        parser.add_argument( '--frame_sync'            , default=False , action='store_true' , help="" )
        parser.add_argument( '--mp4'                   , default=False , action='store_true' , help="" )
        parser.add_argument( '--excel'                 , default=False , action='store_true' , help="" )
        parser.add_argument( '--map_animation'         , default=False , action='store_true' , help="" )
        #############################################################################################

        try:
            ############################################################
            args                                    = parser.parse_args()
            self.__mode                             = int   ( args.mode               )
            self.__json_output_dir                  =         args.json_output_dir
            self.__csv_output_dir                   =         args.csv_output_dir
            self.__movie_output_dir                 =         args.movie_output_dir
            self.__gps_en                           = int   ( args.gps                )
            self.__bme280_en                        = int   ( args.bme280             )
            self.__mpu6050_en                       = int   ( args.mpu6050            )
            self.__icm20948_en                      = int   ( args.icm20948           )
            self.__powermonitor_en                  = int   ( args.powermonitor       )
            self.__icm20948_i2cbus                  = int   ( args.icm20948_i2cbus    )
            self.__bme280_i2cbus                    = int   ( args.bme280_i2cbus      )
            self.__mpu6050_i2cbus                   = int   ( args.mpu6050_i2cbus     )
            self.__framerate                        = int   ( args.framerate          )
            self.__framebuffer                      = int   ( args.framebuffer        )
            self.__gps_interval                     = float ( args.gps_interval       )
            self.__bme280_interval                  = float ( args.bme280_interval    )
            self.__bitrate                          = int   ( args.bitrate            )
            self.__width                            = int   ( args.width              )
            self.__height                           = int   ( args.height             )
            self.__csvbuffer                        = int   ( args.csvbuffer          )
            self.__gps_port                         =         args.gps_port
            self.__bme280_addr                      = int   ( args.bme280_addr   , 16 )
            self.__mpu6050_addr                     = int   ( args.mpu6050_addr  , 16 )
            self.__icm20948_addr                    = int   ( args.icm20948_addr , 16 )
            self.__analyzerDic["input_dir"]         =         args.input_dir
            self.__analyzerDic["frame_sync_en"]     = int   ( args.frame_sync         )
            self.__analyzerDic["mp4_en"]            = int   ( args.mp4                )
            self.__analyzerDic["excel_en"]          = int   ( args.excel              )
            self.__analyzerDic["map_animation_en"]  = int   ( args.map_animation      )
            self.__analyzerDic["calib_json"]        =         args.calib_json
            ############################################################
        except Exception as e:
            print(e)
            sys.exit(1)

    def __setup_sensors( self ):
        print("[Info] Start the __generate_empty_csvFile function.")
        timestamp                  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_output_timestamp_dir   = self.__csv_output_dir   + "/" + timestamp
        movie_output_timestamp_dir = self.__movie_output_dir + "/" + timestamp
        os.makedirs( self.__csv_output_dir      , exist_ok=True )
        os.makedirs( self.__movie_output_dir    , exist_ok=True )
        os.makedirs( csv_output_timestamp_dir   , exist_ok=True )
        os.makedirs( movie_output_timestamp_dir , exist_ok=True )
        cameraCsvFile       = csv_output_timestamp_dir   + "/movie.csv"
        movieFileName       = movie_output_timestamp_dir + "/movie.h264"

        if self.__bme280_i2cbus == self.__mpu6050_i2cbus:
            self.__bme280_bus   = smbus2.SMBus( self.__bme280_i2cbus  )
            self.__mpu6050_bus  = self.__bme280_bus
        else:
            self.__bme280_bus   = smbus2.SMBus( self.__bme280_i2cbus  )
            self.__mpu6050_bus  = smbus2.SMBus( self.__mpu6050_i2cbus )

        #########################################################################
        print("[Info] Activate the Camera Module.")

        data = [
            [
                'start_time'                    ,
                'end_time'                      ,
                'sensor_timestamp'              ,
                'frame_count'                   ,
                'bme280_start_time'             ,
                'bme280_end_time'               ,
                'bme280_byte_0'                 ,
                'bme280_byte_1'                 ,
                'bme280_byte_2'                 ,
                'bme280_byte_3'                 ,
                'bme280_byte_4'                 ,
                'bme280_byte_5'                 ,
                'bme280_byte_6'                 ,
                'bme280_byte_7'                 ,
                'mpu6050_start_time'            ,
                'mpu6050_end_time'              ,
                'mpu6050_byte_0'                ,
                'mpu6050_byte1'                 ,
                'mpu6050_byte2'                 ,
                'mpu6050_byte3'                 ,
                'mpu6050_byte4'                 ,
                'mpu6050_byte5'                 ,
                'mpu6050_byte6'                 ,
                'mpu6050_byte7'                 ,
                'mpu6050_byte8'                 ,
                'mpu6050_byte9'                 ,
                'mpu6050_byte10'                ,
                'mpu6050_byte11'                ,
                'mpu6050_byte12'                ,
                'mpu6050_byte13'                ,
                'icm20948_start_time'           ,
                'icm20948_end_time'             ,
                'icm20948_axRaw'                ,
                'icm20948_ayRaw'                ,
                'icm20948_azRaw'                ,
                'icm20948_gxRaw'                ,
                'icm20948_gyRaw'                ,
                'icm20948_gzRaw'                ,
                'icm20948_mxRaw'                ,
                'icm20948_myRaw'                ,
                'icm20948_mzRaw'                ,
                'icm20948_tmpRaw'               ,
                'ivk172_latitude'               ,
                'ivk172_longitude'              ,
                'ivk172_altitude'               ,
                'ivk172_altitude_units'         ,
                'ivk172_num_sats'               ,
                'ivk172_datestamp'              ,
                'ivk172_timestamp'              ,
                'ivk172_spd_over_grnd'          ,
                'ivk172_true_course'            ,
                'ivk172_true_track'             ,
                'ivk172_spd_over_grnd_kmph'     ,
                'ivk172_pdop'                   ,
                'ivk172_hdop'                   ,
                'ivk172_vdo'                    ,
                'ivk172_num_sv_in_view'         ,
                'ivk172_frame'                  ,
                'powermonitor_start_time'       ,
                'powermonitor_end_time'         ,
                'powermonitor_voltage'          ,
                'powermonitor_throttled'        ,
                'powermonitor_cpu'              ,
                'powermonitor_mem_used_B'       ,
                'powermonitor_mem_total_B'      ,
                'powermonitor_mem_available_B'  ,
                'powermonitor_mem_percent_used' ,
                'powermonitor_temp'             ,
                'powermonitor_disk_used_B'      ,
                'powermonitor_disk_total_B'     ,
                'powermonitor_disk_free_B'      ,
                'powermonitor_disk_percent_used'
            ]
        ]
        self.__generate_empty_csvFile( cameraCsvFile , data )
        self.__camera_fa  = self.__get_csvFile( cameraCsvFile , self.__csvbuffer )

        #######################################################
        # Camera Module Setting
       
        # setting camera configuration
        self.__picamera2     = Picamera2()
        # setting H264 encoder
        encoder = H264Encoder( bitrate=self.__bitrate )
        framerate_microsec   = int( 1.0/self.__framerate*1_000_000 ) # ex) 30fps = 1/30s = 33333μs
        config               = self.__picamera2.create_video_configuration(
            buffer_count = self.__framebuffer                                                              ,
            main         = { "format"             : "YUV420" , "size" : ( self.__width , self.__height ) } ,
            controls     = { "FrameDurationLimits": ( framerate_microsec , framerate_microsec ) }
        )
        self.__picamera2.configure( config )
        ####################################################### 
        
        self.__cameraModuleImpl = CameraModuleImpl(
            self.__picamera2  ,
            encoder           ,
            cameraCsvFile     ,
            self.__camera_fa  ,
            movieFileName
        )
        #########################################################################
        if self.__icm20948_en :
            print("[Info] Activate the ICM-20948.")
            self.__icm20948Impl = ICM20948Impl( self.__icm20948_addr , self.__icm20948_i2cbus )
        #########################################################################
        if self.__bme280_en :
            print("[Info] Activate the BME280.")
            self.__bme280Impl = BME280Impl( self.__bme280_bus , self.__bme280_addr , self.__bme280_interval )
        #########################################################################
        if self.__mpu6050_en:
            print("[Info] Activate the MPU6050.")
            self.__mpu6050Impl = MPU6050Impl( self.__mpu6050_bus , self.__mpu6050_addr )
        #########################################################################
        if self.__gps_en :
            print("[Info] Activate the IVK172 G-Mouse USB GPS.")
            self.__gpsModuleImpl = GPSModuleImpl( self.__gps_port , self.__gps_interval )
        #########################################################################
        if self.__powermonitor_en :
            print("[Info] Activate the PowerMonitor.")
            self.__powermonitorImpl = PowerMonitorImpl()
        #########################################################################

    def doSensorWrapper(self):
        print("[Info] Start the doSensorWrapper function.")
        self.__read_args()
        signal.signal( signal.SIGINT , self.__handler )
        #######################################################################
        if self.__mode == 0:
            print("[Info] It operates in sensor data output mode.")
            self.__setup_sensors()
            threadList = []
            try:
                threadList.append( threading.Thread(                            target=self.__cameraModuleImpl.doCameraModuleImpl ) )
                self.__gps_en          and threadList.append( threading.Thread( target=self.__gpsModuleImpl   .doGpsModuleImpl    ) )
                self.__bme280_en       and threadList.append( threading.Thread( target=self.__bme280Impl      .doBME280Impl       ) )
                self.__mpu6050_en      and threadList.append( threading.Thread( target=self.__mpu6050Impl     .doMPU6050Impl      ) )
                self.__icm20948_en     and threadList.append( threading.Thread( target=self.__icm20948Impl    .doIcm20948Impl     ) )
                self.__powermonitor_en and threadList.append( threading.Thread( target=self.__powermonitorImpl.doPowerMonitorImpl ) )
                SensorWrapper.running.set()
                for singleThread in threadList:
                    singleThread.start()
                for singleThread in threadList:
                    singleThread.join()
            except Exception as e:
                print(e)
        #######################################################################
        elif self.__mode == 1:
            print("[Info] It operates in sensor data reading and analysis mode.")
            sai = SensorAnalyzerImpl( self.__analyzerDic )
            sai.doSensorAnalyzerImpl()
        #######################################################################
        elif self.__mode == 2:
            print("[Info] Start the ICM-20948 calibration mode.")
            cii = CalibrationICM20948Impl( self.__json_output_dir , self.__icm20948_addr , self.__icm20948_i2cbus )
            cii.doCalibrationICM20948Impl()
        #######################################################################

########################################################################
class CalibrationICM20948Impl:

    def __init__(self , output_dir , address , i2cbusnum ):
        self.__output_dir = output_dir
        self.__address    = address
        self.__i2cbusnum  = i2cbusnum
    #######################################################################
    def __save_calibration_to_json(
            self , offset , soft_iron_matrix , accel_range , gyro_range , filename="mag_calibration.json"
    ):
        data = {
            "offset"           : offset.tolist()           ,
            "soft_iron_matrix" : soft_iron_matrix.tolist() ,
            "accel_range"      : accel_range               ,
            "gyro_range"       : gyro_range
        }
        f = open(filename, "w")
        json.dump(data, f, indent=4)
        f.close()                      
    #######################################################################        
    def __collect_mag_samples( self, imu , num_samples=300 ):
        samples = []
        print("[Info] Calibrating... Please move the sensor slowly in all directions.")
        while len(samples) < num_samples:
            imu.getAgmt()
            samples.append([imu.mxRaw, imu.myRaw, imu.mzRaw])
            time.sleep(0.05)
        return numpy.array(samples)
    #######################################################################
    def __compute_offsets( self , samples ):
        mx, my, mz = samples[:, 0], samples[:, 1], samples[:, 2]
        offset_x   = (mx.max() + mx.min()) / 2
        offset_y   = (my.max() + my.min()) / 2
        offset_z   = (mz.max() + mz.min()) / 2
        return numpy.array([offset_x, offset_y, offset_z])
    #######################################################################
    def __compute_soft_iron_matrix( self , centered ):
        cov              = numpy.cov(centered.T)
        eigvals, eigvecs = numpy.linalg.eigh(cov)
        scale            = numpy.diag(1.0 / numpy.sqrt(eigvals))
        soft_iron_matrix = eigvecs @ scale @ eigvecs.T
        return soft_iron_matrix
    #######################################################################
    def __read_accel_range( self , imu ):
        imu.setBank(2)
        reg_val = imu._i2c.readByte( imu.address , 0x14 )
        fs_sel  = (reg_val >> 1) & 0x03
        ranges = {
            0b00: 2  ,
            0b01: 4  ,
            0b10: 8  ,
            0b11: 16
        }
        return ranges.get( fs_sel , 0 )
    #######################################################################
    def __read_gyro_range( self , imu ):
        imu.setBank(2)
        reg_val = imu._i2c.readByte( imu.address , 0x01 )
        fs_sel  = (reg_val >> 1) & 0x03
        ranges = {
            0b00: 250  ,
            0b01: 500  ,
            0b10: 1000 ,
            0b11: 2000
        }
        return ranges.get( fs_sel , 0 )
    #######################################################################
    def doCalibrationICM20948Impl( self ):
        timestamp            = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_timestamp_dir = self.__output_dir + "/" + timestamp
        os.makedirs( self.__output_dir    , exist_ok=True )
        os.makedirs( output_timestamp_dir , exist_ok=True )
        localDriver = qwiic_i2c.getI2CDriver( iBus=self.__i2cbusnum )
        imu         = qwiic_icm20948.QwiicIcm20948( address=self.__address, i2c_driver=localDriver )
        if not imu.connected:
            print("[Error] Could not connect to the ICM-20948.")
            return
        imu.begin()
        samples          = self.__collect_mag_samples(imu)
        offset           = self.__compute_offsets(samples)
        centered         = samples - offset
        soft_iron_matrix = self.__compute_soft_iron_matrix(centered)
        accel_range = self.__read_accel_range(imu)
        gyro_range  = self.__read_gyro_range(imu)
        self.__save_calibration_to_json( offset , soft_iron_matrix , accel_range , gyro_range , output_timestamp_dir+"/mag_calib.json" )

########################################################################
class PowerMonitorImpl:

    def __init__( self ):
        print("[Info] Create an instance of the PowerMonitorImpl class.")
    #######################################################################
    def __get_voltage( self ):
        try:
            out = subprocess.check_output(['vcgencmd', 'measure_volts']).decode()
            return float(out.split('=')[1].replace('V', '').strip())
        except Exception as e:
            print(f"[Error] get_voltage(): {e}")
            return None
    #######################################################################
    def __get_throttled( self ):
        try:
            out = subprocess.check_output(['vcgencmd', 'get_throttled']).decode()
            return out.strip().split('=')[1]
        except Exception as e:
            print(e)
            return None
    #######################################################################
    def __get_memory_usage( self ):
        mem = psutil.virtual_memory()
        return {
            # 'total_MB'     : mem.total     / ( 1024 * 1024 ),
            # 'used_MB'      : mem.used      / ( 1024 * 1024 ),
            # 'available_MB' : mem.available / ( 1024 * 1024 ),
            'total_B'      : mem.total     ,
            'used_B'       : mem.used      ,
            'available_B'  : mem.available ,
            'percent_used' : mem.percent
        }
    #######################################################################
    def __get_cpu_temperature( self ):
        # Bookwormでは /sys/class/thermal の構成が若干異なる場合あり
        try:
            with open("/sys/class/thermal/thermal_zone0/temp", "r") as f:
                #temp = int(f.read()) / 1000  # 単位: 度C
                temp = int(f.read()) # 単位: 度C
            return temp
        except FileNotFoundError:
            return None
    #######################################################################
    def __get_disk_usage( self ):
        disk = psutil.disk_usage('/')
        return {
            # 'total_GB'     : disk.total / ( 1024 ** 3 ),
            # 'used_GB'      : disk.used  / ( 1024 ** 3 ),
            # 'free_GB'      : disk.free  / ( 1024 ** 3 ),
            'total_B'      : disk.total   ,
            'used_B'       : disk.used    ,
            'free_B'       : disk.free    ,
            'percent_used' : disk.percent
        }
    #######################################################################
    def __get_uptime( self ):
        return psutil.boot_time()
    #######################################################################
    def __get_cpu_usage( self ):
        return psutil.cpu_percent(interval=1)
    #######################################################################
    def doPowerMonitorImpl(self):
        print("[Info] Start the doPowerMonitorImpl function.")
        try:
            while SensorWrapper.running.is_set():
                try:
                    SensorWrapper.powermonitor_cond.acquire()
                    while not SensorWrapper.powermonitor_ready:
                        SensorWrapper.powermonitor_cond.wait()
                    SensorWrapper.powermonitor_cond.release()

                    start_time = round( time.monotonic() , 6 )
                    voltage    = self.__get_voltage()
                    throttled  = self.__get_throttled()
                    cpu        = self.__get_cpu_usage()
                    mem        = self.__get_memory_usage()
                    temp       = self.__get_cpu_temperature()
                    disk       = self.__get_disk_usage()
                    
                    SensorWrapper.powermonitor_start_time        = start_time
                    SensorWrapper.powermonitor_end_time          = round( time.monotonic() , 6 )
                    SensorWrapper.powermonitor_voltage           = voltage
                    SensorWrapper.powermonitor_throttled         = throttled
                    SensorWrapper.powermonitor_cpu               = cpu
                    SensorWrapper.powermonitor_mem_used_B        = mem['used_B']
                    SensorWrapper.powermonitor_mem_total_B       = mem['total_B']
                    SensorWrapper.powermonitor_mem_available_B   = mem['available_B']
                    SensorWrapper.powermonitor_mem_percent_used  = mem['percent_used']
                    SensorWrapper.powermonitor_temp              = temp
                    SensorWrapper.powermonitor_disk_used_B       = disk['used_B']
                    SensorWrapper.powermonitor_disk_total_B      = disk['total_B']
                    SensorWrapper.powermonitor_disk_free_B       = disk['free_B']
                    SensorWrapper.powermonitor_disk_percent_used = disk['percent_used']

                    SensorWrapper.powermonitor_ready = False
                except (KeyboardInterrupt , ValueError) as e:
                    SensorWrapper.running.clear()
                except Exception as e:
                    print(e)
        finally:
            stop_event.set()
            writer_thread.join()

########################################################################
class BME280Impl:

    def __init__( self , bus , address , interval ):
        print("[Info] Create an instance of the BME280Impl class.")
        print("[Info] The device address of the BME280 is " + str(hex(address)) )
        self.__address       = address
        self.__bus           = bus
        self.__interval      = interval
        self.__bus.write_byte_data( self.__address , 0xF2 , 0x01 )  # Humidity oversampling x1
        self.__bus.write_byte_data( self.__address , 0xF4 , 0x27 )  # Normal mode, temp/press oversampling x1
        self.__bus.write_byte_data( self.__address , 0xF5 , 0xA0 )  # Config
    #######################################################################
    def __read_sensor(self):
        # first only
        # read24byte    = self.__bus.read_i2c_block_data ( self.__address , 0x88 , 24 ) # calib
        # read1Byte0xA1 = self.__bus.read_byte_data      ( self.__address , 0xA1      ) # calib
        # read7byte     = self.__bus.read_i2c_block_data ( self.__address , 0xE1 ,  7 ) # calib
        read8byte     = self.__bus.read_i2c_block_data ( self.__address , 0xF7 ,  8 ) # 
        return read8byte
    #######################################################################
    def doBME280Impl(self):
        print("[Info] Start the doBME280Impl function.")
        #try:
        while SensorWrapper.running.is_set():
                # try:
                #     SensorWrapper.bme280_cond.acquire()
                #     while not SensorWrapper.bme280_ready:
                #         SensorWrapper.bme280_cond.wait()
                #     SensorWrapper.bme280_cond.release()

            start_time = round( time.monotonic() , 6 )
            read8byte  = self.__read_sensor()

            SensorWrapper.bme280_start_time = start_time
            SensorWrapper.bme280_end_time   = round( time.monotonic() , 6)
            SensorWrapper.bme280_byte_0     = read8byte[0]
            SensorWrapper.bme280_byte_1     = read8byte[1]
            SensorWrapper.bme280_byte_2     = read8byte[2]
            SensorWrapper.bme280_byte_3     = read8byte[3]   
            SensorWrapper.bme280_byte_4     = read8byte[4]
            SensorWrapper.bme280_byte_5     = read8byte[5]
            SensorWrapper.bme280_byte_6     = read8byte[6]
            SensorWrapper.bme280_byte_7     = read8byte[7]

            time.sleep(self.__interval)
                    
                #     SensorWrapper.bme280_ready = False
                # except (KeyboardInterrupt , ValueError) as e:
                #     SensorWrapper.running.clear()
                # except Exception as e:
                #     print(e)
        # finally:
        #     stop_event.set()
        #     writer_thread.join()

########################################################################
class MPU6050Impl:

    def __init__( self , bus , address ):
        print("[Info] Create an instance of the MPU6050Impl class.")
        print("[Info] The device address of the MPU6050 is " + str(hex(address)) )
        self.__address       = address
        self.__bus           = bus
        self.__bus.write_byte_data( self.__address , 0x6B , 0x00 ) # スリープ解除
        self.__bus.write_byte_data( self.__address , 0x1A , 0x03 ) # DLPF設定
        self.__bus.write_byte_data( self.__address , 0x1B , 0x00 ) # ジャイロフルスケール
        self.__bus.write_byte_data( self.__address , 0x1C , 0x00 ) # 加速度フルスケール
    #######################################################################
    def __read_sensor( self ):
        try:
            mpu6050_data = self.__bus.read_i2c_block_data( self.__address , 0x3B , 14 )
            return mpu6050_data
        except:
            return None
    #######################################################################
    def doMPU6050Impl(self):
        print("[Info] Start the doMPU6050Impl function.")
        try:
            while SensorWrapper.running.is_set():
                try:
                    SensorWrapper.mpu6050_cond.acquire()
                    while not SensorWrapper.mpu6050_ready:
                        SensorWrapper.mpu6050_cond.wait()
                    SensorWrapper.mpu6050_cond.release()

                    start_time = round( time.monotonic() , 6 )
                    mpu6050_data = None
                    retry_count  = 0
                    while ((mpu6050_data is None) or (retry_count==10)):
                        mpu6050_data = self.__read_sensor()
                        if mpu6050_data is None:
                            # retry
                            time.sleep(0.001)
                            retry_count += 1

                    if mpu6050_data is not None:
                        SensorWrapper.mpu6050_start_time = start_time
                        SensorWrapper.mpu6050_end_time   = round( time.monotonic() , 6 )
                        SensorWrapper.mpu6050_byte_0     = mpu6050_data[ 0]
                        SensorWrapper.mpu6050_byte1      = mpu6050_data[ 1]
                        SensorWrapper.mpu6050_byte2      = mpu6050_data[ 2] 
                        SensorWrapper.mpu6050_byte3      = mpu6050_data[ 3]
                        SensorWrapper.mpu6050_byte4      = mpu6050_data[ 4]
                        SensorWrapper.mpu6050_byte5      = mpu6050_data[ 5] 
                        SensorWrapper.mpu6050_byte6      = mpu6050_data[ 6]
                        SensorWrapper.mpu6050_byte7      = mpu6050_data[ 7]
                        SensorWrapper.mpu6050_byte8      = mpu6050_data[ 8] 
                        SensorWrapper.mpu6050_byte9      = mpu6050_data[ 9]
                        SensorWrapper.mpu6050_byte10     = mpu6050_data[10]
                        SensorWrapper.mpu6050_byte11     = mpu6050_data[11] 
                        SensorWrapper.mpu6050_byte12     = mpu6050_data[12]
                        SensorWrapper.mpu6050_byte13     = mpu6050_data[13]
                        
                    SensorWrapper.mpu6050_ready = False
                except (KeyboardInterrupt , ValueError) as e:
                    SensorWrapper.running.clear()
                except Exception as e:
                    print(e)
        finally:
            stop_event.set()
            writer_thread.join()

########################################################################
class ICM20948Impl:

    def __init__( self , address , i2cbus ):
        print("[Info] Create an instance of the ICM20948Impl class.")
        print("[Info] The device address of the ICM-20948 is " + str(hex(address)) )
        localDriver = qwiic_i2c.getI2CDriver( iBus=i2cbus )
        self.__imu  = qwiic_icm20948.QwiicIcm20948( address=address, i2c_driver=localDriver )
    #######################################################################
    def doIcm20948Impl(self):
        print("[Info] Start the doIcm20948Impl function.")
        try:
            self.__imu.begin()
            while SensorWrapper.running.is_set():
                try:
                    SensorWrapper.icm20948_cond.acquire()
                    while not SensorWrapper.icm20948_ready:
                        SensorWrapper.icm20948_cond.wait()
                    SensorWrapper.icm20948_cond.release()

                    start_time = round( time.monotonic() , 6 )
                    if self.__imu.dataReady():
                        self.__imu.getAgmt()

                        SensorWrapper.icm20948_start_time = start_time
                        SensorWrapper.icm20948_end_time   = round( time.monotonic() , 6 )
                        SensorWrapper.icm20948_axRaw      = self.__imu.axRaw
                        SensorWrapper.icm20948_ayRaw      = self.__imu.ayRaw
                        SensorWrapper.icm20948_azRaw      = self.__imu.azRaw 
                        SensorWrapper.icm20948_gxRaw      = self.__imu.gxRaw
                        SensorWrapper.icm20948_gyRaw      = self.__imu.gyRaw
                        SensorWrapper.icm20948_gzRaw      = self.__imu.gzRaw 
                        SensorWrapper.icm20948_mxRaw      = self.__imu.mxRaw
                        SensorWrapper.icm20948_myRaw      = self.__imu.myRaw
                        SensorWrapper.icm20948_mzRaw      = self.__imu.mzRaw 
                        SensorWrapper.icm20948_tmpRaw     = self.__imu.tmpRaw

                        SensorWrapper.icm20948_ready = False
                except (KeyboardInterrupt , ValueError) as e:
                    SensorWrapper.running.clear()
                except Exception as e:
                    print(e)
        finally:
            stop_event.set()
            writer_thread.join()        
                    
########################################################################
class GPSModuleImpl:

    def __init__( self , port , interval ):
        print("[Info] Create an instance of the GPSModuleImpl class.")
        print("[Info] The port for the IVK172 G-Mouse USB GPS is " + str(port) + ".")
        self.__ser           = serial.Serial( port , 9600 , timeout=1 )
        self.__interval      = interval
    #######################################################################
    def __read_sensor( self ):
        frame = {"GGA": None, "RMC": None, "VTG": None, "GSA": None, "GSV": None}
        try:
            while True:
                start_time = round( time.monotonic() , 6 )
                raw = self.__ser.readline().decode('ascii', errors='replace').strip()
                if not raw.startswith('$'):
                    continue
                msg = None
                try:
                    rawmsg = pynmea2.parse(raw)
                    msg = rawmsg
                except pynmea2.ParseError:
                    pass
                if msg is None:
                    continue
                key = msg.sentence_type
                if key in frame:
                    frame[key] = msg
                if all(frame.values()):
                    gga , rmc , vtg , gsa , gsv = (
                        frame["GGA"] , frame["RMC"] , frame["VTG"] , frame["GSA"] , frame["GSV"]
                    )

                    SensorWrapper.ivk172_start_time         = start_time
                    SensorWrapper.ivk172_latitude           = gga.latitude
                    SensorWrapper.ivk172_longitude          = gga.longitude
                    SensorWrapper.ivk172_altitude           = gga.altitude
                    SensorWrapper.ivk172_altitude_units     = gga.altitude_units
                    SensorWrapper.ivk172_num_sats           = gga.num_sats
                    SensorWrapper.ivk172_datestamp          = rmc.datestamp
                    SensorWrapper.ivk172_timestamp          = rmc.timestamp
                    SensorWrapper.ivk172_spd_over_grnd      = rmc.spd_over_grnd
                    SensorWrapper.ivk172_true_course        = rmc.true_course
                    SensorWrapper.ivk172_true_track         = vtg.true_track
                    SensorWrapper.ivk172_spd_over_grnd_kmph = vtg.spd_over_grnd_kmph
                    SensorWrapper.ivk172_pdop               = gsa.pdop
                    SensorWrapper.ivk172_hdop               = gsa.hdop
                    SensorWrapper.ivk172_vdo                = gsa.vdop
                    SensorWrapper.ivk172_num_sv_in_view     = gsv.num_sv_in_view
                    SensorWrapper.ivk172_frame              = dict.fromkeys( frame , None )
                    SensorWrapper.ivk172_end_time           = round( time.monotonic() , 6 )
                    
                time.sleep( self.__interval )
        except KeyboardInterrupt as e:
            pass # ignore
        finally:
            self.__ser.close()
    #######################################################################
    def doGpsModuleImpl(self):
        print("[Info] Start the doGpsModuleImpl function.")
        try:
            self.__read_sensor()
        except Exception as e:
            pass # ignore

########################################################################
class CameraModuleImpl:

    write_queue = queue.Queue()

    def __init__( self , picamera2 , encoder , csvFileName , csvFile , movieFileName ):
        print("[Info] Create an instance of the CameraModuleImpl class.")
        self.__frame_ready             = threading.Event()
        self.__frame_count             = 0
        self.__encoder                 = encoder
        self.__picamera2               = picamera2
        self.__picamera2.post_callback = self.__process_frame
        self.__movieFile               = movieFileName
        self.__csvFile                 = csvFile
        self.__csvFileWriter           = csv.writer( csvFile )
        self.__end_time                = None
        self.__sensor_ts               = None
    #######################################################################
    def __csv_writer( self , stop_event ):
        while not stop_event.is_set() or not CameraModuleImpl.write_queue.empty():
            try:
                row = CameraModuleImpl.write_queue.get( timeout=0.1 )
                self.__csvFileWriter.writerow( row )
            except queue.Empty:
                continue
        self.__csvFile.flush()
    #######################################################################
    def __process_frame( self, request ):
        SensorWrapper.camera_module_cond    .acquire()
        #SensorWrapper.bme280_cond           .acquire()
        SensorWrapper.mpu6050_cond          .acquire()
        SensorWrapper.icm20948_cond         .acquire()
        SensorWrapper.powermonitor_cond     .acquire()
        if ((self.__frame_count % 30) == 0) :
            #SensorWrapper.bme280_cond      .notify()
            SensorWrapper.powermonitor_cond.notify()      
        SensorWrapper.mpu6050_cond         .notify()
        SensorWrapper.icm20948_cond        .notify()
        SensorWrapper.camera_module_cond   .notify()
        # SensorWrapper.bme280_ready        = True
        SensorWrapper.mpu6050_ready       = True
        SensorWrapper.icm20948_ready      = True
        SensorWrapper.powermonitor_ready  = True
        SensorWrapper.camera_module_ready = True
        # SensorWrapper.bme280_cond          .release()
        SensorWrapper.mpu6050_cond         .release()
        SensorWrapper.icm20948_cond        .release()
        SensorWrapper.powermonitor_cond    .release()
        SensorWrapper.camera_module_cond   .release()
        self.__sensor_ts = request.get_metadata().get( "SensorTimestamp" , 0 )
        self.__end_time  = round( time.monotonic() , 6 )
        self.__frame_ready.set()
    #######################################################################
    def __output_camera_module_csv( self ):
        stop_event    = threading.Event()
        writer_thread = threading.Thread( target=self.__csv_writer , args=( stop_event, ) )
        writer_thread.start()
        try:
            while SensorWrapper.running.is_set():
                try:
                    SensorWrapper.camera_module_cond.acquire()
                    while not SensorWrapper.camera_module_ready:
                        SensorWrapper.camera_module_cond.wait()
                    SensorWrapper.camera_module_cond.release()
    
                    data = [
                        SensorWrapper.start_time                    ,
                        self.__end_time                             ,
                        self.__sensor_ts                            ,
                        self.__frame_count                          ,
                        SensorWrapper.bme280_start_time             ,
                        SensorWrapper.bme280_end_time               ,
                        SensorWrapper.bme280_byte_0                 ,
                        SensorWrapper.bme280_byte_1                 ,
                        SensorWrapper.bme280_byte_2                 ,
                        SensorWrapper.bme280_byte_3                 ,
                        SensorWrapper.bme280_byte_4                 ,
                        SensorWrapper.bme280_byte_5                 ,
                        SensorWrapper.bme280_byte_6                 ,
                        SensorWrapper.bme280_byte_7                 ,
                        SensorWrapper.mpu6050_start_time            ,
                        SensorWrapper.mpu6050_end_time              ,
                        SensorWrapper.mpu6050_byte_0                ,
                        SensorWrapper.mpu6050_byte1                 ,
                        SensorWrapper.mpu6050_byte2                 ,
                        SensorWrapper.mpu6050_byte3                 ,
                        SensorWrapper.mpu6050_byte4                 ,
                        SensorWrapper.mpu6050_byte5                 ,
                        SensorWrapper.mpu6050_byte6                 ,
                        SensorWrapper.mpu6050_byte7                 ,
                        SensorWrapper.mpu6050_byte8                 ,
                        SensorWrapper.mpu6050_byte9                 ,
                        SensorWrapper.mpu6050_byte10                ,
                        SensorWrapper.mpu6050_byte11                ,
                        SensorWrapper.mpu6050_byte12                ,
                        SensorWrapper.mpu6050_byte13                ,
                        SensorWrapper.icm20948_start_time           ,
                        SensorWrapper.icm20948_end_time             ,
                        SensorWrapper.icm20948_axRaw                ,
                        SensorWrapper.icm20948_ayRaw                ,
                        SensorWrapper.icm20948_azRaw                ,
                        SensorWrapper.icm20948_gxRaw                ,
                        SensorWrapper.icm20948_gyRaw                ,
                        SensorWrapper.icm20948_gzRaw                ,
                        SensorWrapper.icm20948_mxRaw                ,
                        SensorWrapper.icm20948_myRaw                ,
                        SensorWrapper.icm20948_mzRaw                ,
                        SensorWrapper.icm20948_tmpRaw               ,
                        SensorWrapper.ivk172_latitude               ,
                        SensorWrapper.ivk172_longitude              ,
                        SensorWrapper.ivk172_altitude               ,
                        SensorWrapper.ivk172_altitude_units         ,
                        SensorWrapper.ivk172_num_sats               ,
                        SensorWrapper.ivk172_datestamp              ,
                        SensorWrapper.ivk172_timestamp              ,
                        SensorWrapper.ivk172_spd_over_grnd          ,
                        SensorWrapper.ivk172_true_course            ,
                        SensorWrapper.ivk172_true_track             ,
                        SensorWrapper.ivk172_spd_over_grnd_kmph     ,
                        SensorWrapper.ivk172_pdop                   ,
                        SensorWrapper.ivk172_hdop                   ,
                        SensorWrapper.ivk172_vdo                    ,
                        SensorWrapper.ivk172_num_sv_in_view         ,
                        SensorWrapper.ivk172_frame                  ,
                        SensorWrapper.powermonitor_start_time       ,
                        SensorWrapper.powermonitor_end_time         ,
                        SensorWrapper.powermonitor_voltage          ,
                        SensorWrapper.powermonitor_throttled        ,
                        SensorWrapper.powermonitor_cpu              ,
                        SensorWrapper.powermonitor_mem_used_B       ,
                        SensorWrapper.powermonitor_mem_total_B      ,
                        SensorWrapper.powermonitor_mem_available_B  ,
                        SensorWrapper.powermonitor_mem_percent_used ,
                        SensorWrapper.powermonitor_temp             ,
                        SensorWrapper.powermonitor_disk_used_B      ,
                        SensorWrapper.powermonitor_disk_total_B     ,
                        SensorWrapper.powermonitor_disk_free_B      ,
                        SensorWrapper.powermonitor_disk_percent_used
                    ]
                    
                    CameraModuleImpl.write_queue.put( data )
                    self.__frame_count += 1
                    SensorWrapper.camera_module_ready = False
                except (KeyboardInterrupt , ValueError) as e:
                     SensorWrapper.running.clear()
                except Exception as e:
                    print(e)
        finally:
            stop_event.set()
            writer_thread.join()
    #######################################################################
    def doCameraModuleImpl( self ):
        print("[Info] Start the doCameraModuleImpl function.")
        SensorWrapper.start_time = round( time.monotonic() , 6 )

        cameraThread = threading.Thread( target=self.__output_camera_module_csv )
        cameraThread.start()

        self.__picamera2.start()
        self.__picamera2.start_encoder( self.__encoder , output=self.__movieFile )

        while SensorWrapper.running.is_set():
            try:
                if self.__frame_ready.wait(timeout=1.0):
                    self.__frame_ready.clear()
                cameraThread.join()
            except (KeyboardInterrupt , ValueError) as e:
                SensorWrapper.running.clear()
            except Exception as e:
                print(e)
            finally:
                self.__picamera2.stop_encoder()
                self.__picamera2.stop()

#############################################################################################################
#############################################################################################################
#############################################################################################################
class SensorAnalyzerImpl:

    def __init__( self , analyzerDic ):
        print("[Info] Create an instance of the SensorAnalyzerImpl class.")
        self.__parameterDic = analyzerDic

    #########################################################################
    def doSensorAnalyzerImpl( self ):
        print("[Info] Start the doSensorAnalyzerImpl function.")
        try:
            threadList = []

            # GPSデータが存在する場合
            if os.path.isfile( self.__parameterDic["input_dir"] + "/" + "gps.csv" ):
                gai = GPSAnalyzerImpl(
                    self.__parameterDic["input_dir"] + "/" + "gps.csv" ,
                    self.__parameterDic["map_animation_en"]
                )
                # GPSデータを地図、GoogleMapデータで可視化できるようにする
                threadList.append( threading.Thread( target=gai.doGPSAnalyzerImpl ) )

            # 動画データが存在する場合
            if os.path.isfile( self.__parameterDic["input_dir"] + "/" + "movie.h264" ):
                iai = I2CAnalyzerImpl(
                    self.__parameterDic["input_dir"]  ,
                    self.__parameterDic["excel_en"]   ,
                    self.__parameterDic["calib_json"]
                )
                mai = MovieAnalyzerImpl( self.__parameterDic["input_dir"] )
                
                # MP4に変換したい場合
                if self.__parameterDic["mp4_en"]:
                    threadList.append(
                        threading.Thread(
                            target = mai.doMovieAnalyzerImpl(
                                False ,
                                self.__parameterDic["input_dir"] + "/" + "movie.h264" ,
                                None
                            )
                        )
                    )

                # 動画とセンサーデータを同期したい場合
                #  動画のcsvファイルが存在する場合実行
                if (
                        self.__parameterDic["frame_sync_en"]                                   and
                        os.path.isfile( self.__parameterDic["input_dir"] + "/" + "movie.csv" ) and
                        (
                            os.path.isfile( self.__parameterDic["input_dir"] + "/" + "bme280.csv"       ) or
                            os.path.isfile( self.__parameterDic["input_dir"] + "/" + "icm20948.csv"     ) or
                            os.path.isfile( self.__parameterDic["input_dir"] + "/" + "mpu6050.csv"      ) or
                            os.path.isfile( self.__parameterDic["input_dir"] + "/" + "powermonitor.csv" )
                        )
                ):
                    # 動画のcsvと他センサーデータのcsvをマージする
                    # マージ後データを動画データに組み込む
                    iai.doI2CAnalyzerImpl()
                    threadList.append(
                        threading.Thread(
                            target= mai.doMovieAnalyzerImpl(
                                self.__parameterDic["frame_sync_en"] ,
                                self.__parameterDic["input_dir"] + "/" + "movie.h264" ,
                                iai.getMergeDataFrame()
                            )
                        )
                    )

            for signleThread in threadList:
                signleThread.start()
            for signleThread in threadList:
                signleThread.join()

        except Exception as e:
            print(e)

########################################################################################
class I2CAnalyzerImpl:

    def __init__( self , input_dir , excel_en , calib_json ):
        self.__input_dir        = input_dir
        self.__excel_en         = excel_en
        self.__calib_json       = calib_json
        self.__camera_csv       = input_dir + "/" + "movie.csv"
        self.__gps_csv          = input_dir + "/" + "gps.csv"
        self.__powermonitor_csv = input_dir + "/" + "powermonitor.csv"
        self.__bme280_csv       = input_dir + "/" + "bme280.csv"
        self.__icm20948_csv     = input_dir + "/" + "icm20948.csv"
        self.__mpu6050_csv      = input_dir + "/" + "mpu6050.csv"
        self.__mergeDataFrame   = None

    def __merge_csv( self ):
        print("[Info] Start the __merge_csv function.")
        timestamp_column  = "end_unix_epoch_time"
        base_csv_filename = "movie.csv"
        tolerance_sec     = 100 # 100sec
        csv_files         = sorted( glob.glob(os.path.join(self.__input_dir, "*.csv")) )
        base_csv_path     = os.path.join( self.__input_dir, base_csv_filename )
        dfs = {}
        for file in csv_files:
            df       = pandas.read_csv(file)
            basename = os.path.splitext(os.path.basename(file))[0]
            df[f"{basename}_end_unix_epoch_time"] = df[timestamp_column]
            # ミリ秒-> 秒判定と変換
            if df[timestamp_column].max() > 1e12:
                df[timestamp_column] = df[timestamp_column] / 1000.0
            df[timestamp_column] = pandas.to_datetime(df[timestamp_column], unit='s')
            dfs[file] = df
        base_df = dfs[base_csv_path].copy()
        base_df = base_df.sort_values(by=timestamp_column).reset_index(drop=True)

        for file, other_df in dfs.items():
            if file == base_csv_path:
                continue
            basename      = os.path.splitext(os.path.basename(file))[0]
            epoch_col     = f"{basename}_end_unix_epoch_time"
            temp_time_col = f"_temp_time_{basename}"
            other_df      = other_df.rename(columns={timestamp_column: temp_time_col})
            other_df      = other_df.sort_values(by=temp_time_col).reset_index(drop=True)

            # マージ(toleranceあり)
            merged = pandas.merge_asof(
                base_df                      ,
                other_df                     ,
                left_on   = timestamp_column ,
                right_on  = temp_time_col    ,
                direction = "nearest"        ,
                tolerance = pandas.Timedelta(seconds=tolerance_sec)
            )

            # 補完処理
            missing_mask = merged.filter(like='_y').isnull().any(axis=1)
            if missing_mask.any():
                fallback = pandas.merge_asof(
                    base_df[missing_mask]       ,
                    other_df                    ,
                    left_on   =timestamp_column ,
                    right_on  =temp_time_col    ,
                    direction ="nearest"
                )
                merged.loc[missing_mask] = fallback

            merged  = merged.drop(columns=[temp_time_col])
            base_df = merged

        base_df     = base_df.drop(columns=[timestamp_column])
        output_path = f"merged_cleaned_{os.path.splitext(base_csv_filename)[0]}.csv"
        #base_df.to_csv(output_path, index=False)
        self.__mergeDataFrame = base_df

    def getMergeDataFrame( self ):
        return self.__mergeDataFrame

    #########################################################################
    def __output_to_excel( self , sheetName , fileName , dataFrame ):
        print("[Info] Start the __output_to_excel function.")
        wb              = openpyxl.Workbook()
        ws              = wb.active
        ws.title        = sheetName
        ws.freeze_panes = 'B2'
        font            = Font(name='BIZ UDゴシック', size=11)
        no_border = Border(
            # left   = Side( border_style=None ) ,
            # right  = Side( border_style=None ) ,
            # top    = Side( border_style=None ) ,
            # bottom = Side( border_style=None )
            left    = Side( style='thin' ) ,
            right   = Side( style='thin' ) ,
            top     = Side( style='thin' ) ,
            bottom  = Side( style='thin' )
        )

        # Data writing & font/border settings
        for r_idx, row in enumerate( dataframe_to_rows( dataFrame , index=False , header=True ) , start=1 ):
            for c_idx , value in enumerate( row , start=1 ):
                cell        = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font   = font
                cell.border = no_border
                if r_idx == 1:
                    cell.fill      = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    cell.alignment = Alignment( horizontal='center', vertical='center' )

        # Auto-adjust column width
        for col_idx, column_cells in enumerate( ws.columns , start=1 ):
            max_length     = max( len( str(cell.value) ) if cell.value is not None else 0 for cell in column_cells )
            adjusted_width = max_length + 2.5
            col_letter     = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted_width
        ws.auto_filter.ref = ws.dimensions
        wb.save( fileName )

    ##################################################################################
    # BME280
    ##################################################################################
    def __convert_bme280_dataFrame( self ):
        print("[Info] Start the __convert_bme280_dataFrame function.")
        # センサーレジスタデータを物理量データに変換
        self.__mergeDataFrame[
            [
                'bme280_temperature' ,
                'bme280_pressure'    ,
                'bme280_humidity'    ,
                'bme280_altitude'
            ]
        ] = self.__mergeDataFrame.apply(
            lambda row: pandas.Series(
                self.__convert_bme280_batch(
                    row['bme280_byte_00'] , row['bme280_byte_01'] , row['bme280_byte_02'] ,
                    row['bme280_byte_03'] , row['bme280_byte_04'] , row['bme280_byte_05'] ,
                    row['bme280_byte_06'] , row['bme280_byte_07'] , row['bme280_byte_08'] ,
                    row['bme280_byte_09'] , row['bme280_byte_10'] , row['bme280_byte_11'] ,
                    row['bme280_byte_12'] , row['bme280_byte_13'] , row['bme280_byte_14'] ,
                    row['bme280_byte_15'] , row['bme280_byte_16'] , row['bme280_byte_17'] ,
                    row['bme280_byte_18'] , row['bme280_byte_19'] , row['bme280_byte_20'] ,
                    row['bme280_byte_21'] , row['bme280_byte_22'] , row['bme280_byte_23'] ,
                    row['bme280_byte_24'] , row['bme280_byte_25'] , row['bme280_byte_26'] ,
                    row['bme280_byte_27'] , row['bme280_byte_28'] , row['bme280_byte_29'] ,
                    row['bme280_byte_30'] , row['bme280_byte_31'] , row['bme280_byte_32'] ,
                    row['bme280_byte_33'] , row['bme280_byte_34'] , row['bme280_byte_35'] ,
                    row['bme280_byte_36'] , row['bme280_byte_37'] , row['bme280_byte_38'] ,
                    row['bme280_byte_39']
                )
            ) , axis=1
        )
        self.__mergeDataFrame = self.__mergeDataFrame.drop(
            [
                'bme280_byte_00' , 'bme280_byte_01' , 'bme280_byte_02' ,
                'bme280_byte_03' , 'bme280_byte_04' , 'bme280_byte_05' ,
                'bme280_byte_06' , 'bme280_byte_07' , 'bme280_byte_08' ,
                'bme280_byte_09' , 'bme280_byte_10' , 'bme280_byte_11' ,
                'bme280_byte_12' , 'bme280_byte_13' , 'bme280_byte_14' ,
                'bme280_byte_15' , 'bme280_byte_16' , 'bme280_byte_17' ,
                'bme280_byte_18' , 'bme280_byte_19' , 'bme280_byte_20' ,
                'bme280_byte_21' , 'bme280_byte_22' , 'bme280_byte_23' ,
                'bme280_byte_24' , 'bme280_byte_25' , 'bme280_byte_26' ,
                'bme280_byte_27' , 'bme280_byte_28' , 'bme280_byte_29' ,
                'bme280_byte_30' , 'bme280_byte_31' , 'bme280_byte_32' ,
                'bme280_byte_33' , 'bme280_byte_34' , 'bme280_byte_35' ,
                'bme280_byte_36' , 'bme280_byte_37' , 'bme280_byte_38' ,
                'bme280_byte_39'
            ] ,
            axis=1
        )
    ##############################################################################
    def __convert_bme280_batch(
            self       ,
            byteData00 , byteData01 , byteData02 , byteData03 , byteData04 , byteData05 , byteData06 ,
            byteData07 , byteData08 , byteData09 , byteData10 , byteData11 , byteData12 , byteData13 ,
            byteData14 , byteData15 , byteData16 , byteData17 , byteData18 , byteData19 , byteData20 ,
            byteData21 , byteData22 , byteData23 , byteData24 , byteData25 , byteData26 , byteData27 ,
            byteData28 , byteData29 , byteData30 , byteData31 , byteData32 , byteData33 , byteData34 ,
            byteData35 , byteData36 , byteData37 , byteData38 , byteData39
    ):
        dig_T1 = int(byteData01) << 8 | int(byteData00)
        dig_T2 = ( int(byteData03) << 8 | int(byteData02) ) if int(byteData03) < 128 else ( int(byteData03) << 8 | int(byteData02) ) - 65536
        dig_T3 = ( int(byteData05) << 8 | int(byteData04) ) if int(byteData05) < 128 else ( int(byteData05) << 8 | int(byteData04) ) - 65536
        dig_P1 = int(byteData07) << 8 | int(byteData06)
        dig_P2 = ( int(byteData09) << 8 | int(byteData08) ) if int(byteData09) < 128 else ( int(byteData09) << 8 | int(byteData08) ) - 65536
        dig_P3 = ( int(byteData11) << 8 | int(byteData10) ) if int(byteData11) < 128 else ( int(byteData11) << 8 | int(byteData10) ) - 65536
        dig_P4 = ( int(byteData13) << 8 | int(byteData12) ) if int(byteData13) < 128 else ( int(byteData13) << 8 | int(byteData12) ) - 65536
        dig_P5 = ( int(byteData15) << 8 | int(byteData14) ) if int(byteData15) < 128 else ( int(byteData15) << 8 | int(byteData14) ) - 65536
        dig_P6 = ( int(byteData17) << 8 | int(byteData16) ) if int(byteData17) < 128 else ( int(byteData17) << 8 | int(byteData16) ) - 65536
        dig_P7 = ( int(byteData19) << 8 | int(byteData18) ) if int(byteData19) < 128 else ( int(byteData19) << 8 | int(byteData18) ) - 65536
        dig_P8 = ( int(byteData21) << 8 | int(byteData20) ) if int(byteData21) < 128 else ( int(byteData21) << 8 | int(byteData20) ) - 65536
        dig_P9 = ( int(byteData23) << 8 | int(byteData22) ) if int(byteData23) < 128 else ( int(byteData23) << 8 | int(byteData22) ) - 65536
        dig_H1 = int(byteData24)
        dig_H2 = ( int(byteData26) << 8 | int(byteData25) ) if int(byteData26) < 128 else ( int(byteData26) << 8 | int(byteData25) ) - 65536
        dig_H3 = int(byteData27)
        dig_H4 = ( int(byteData28) << 4 ) | ( int(byteData29) & 0x0F )
        if dig_H4 & 0x800: dig_H4 -= 4096
        dig_H5 = ( int(byteData30) << 4 ) | ( int(byteData29) >> 4 )
        if dig_H5 & 0x800: dig_H5 -= 4096
        dig_H6 = int(byteData31)
        if dig_H6 > 127: dig_H6 -= 256
        calib = {
            'T': ( dig_T1 , dig_T2 , dig_T3 ) ,
            'P': ( dig_P1 , dig_P2 , dig_P3 , dig_P4 , dig_P5 , dig_P6 , dig_P7 , dig_P8 , dig_P9 ) ,
            'H': ( dig_H1 , dig_H2 , dig_H3 , dig_H4 , dig_H5 , dig_H6 )
        }
        adc_P = ( int(byteData32) << 12 ) | ( int(byteData33) << 4 ) | ( int(byteData34) >> 4 )
        adc_T = ( int(byteData35) << 12 ) | ( int(byteData36) << 4 ) | ( int(byteData37) >> 4 )
        adc_H = ( int(byteData38) <<  8 ) |   int(byteData39)
        temp, t_fine = self.__compensate_temperature( adc_T , calib['T']          )
        press        = self.__compensate_pressure   ( adc_P , calib['P'] , t_fine )
        hum          = self.__compensate_humidity   ( adc_H , calib['H'] , t_fine )
        alti         = self.__calculate_altitude( temp , hum , press )
        return temp , press , hum , alti
    ##############################################################################
    def __compensate_temperature( self , adc_T , calib ):
        dig_T1, dig_T2, dig_T3 = calib
        var1 = (((adc_T >> 3) - (dig_T1 << 1)) * dig_T2) >> 11
        var2 = (((((adc_T >> 4) - dig_T1) * ((adc_T >> 4) - dig_T1)) >> 12) * dig_T3) >> 14
        t_fine = var1 + var2
        temperature = (t_fine * 5 + 128) >> 8
        return temperature / 100.0, t_fine
    ##############################################################################
    def __compensate_pressure( self , adc_p, dig_P, t_fine):
        var1 = t_fine / 2.0 - 64000.0
        var2 = var1 * var1 * dig_P[5] / 32768.0
        var2 = var2 + var1 * dig_P[4] * 2.0
        var2 = var2 / 4.0 + dig_P[3] * 65536.0
        var1 = (dig_P[2] * var1 * var1 / 524288.0 + dig_P[1] * var1) / 524288.0
        var1 = (1.0 + var1 / 32768.0) * dig_P[0]
        if var1 == 0:
            return 0
        p = 1048576.0 - adc_p
        p = ((p - var2 / 4096.0) * 6250.0) / var1
        var1 = dig_P[8] * p * p / 2147483648.0
        var2 = p * dig_P[7] / 32768.0
        p = p + (var1 + var2 + dig_P[6]) / 16.0
        return p / 100.0
    ##############################################################################
    def __compensate_humidity( self , adc_H , calib , t_fine ):
        dig_H1, dig_H2, dig_H3, dig_H4, dig_H5, dig_H6 = calib
        v_x1 = t_fine - 76800
        v_x1 = (((((adc_H << 14) - (dig_H4 << 20) - (dig_H5 * v_x1)) + 16384) >> 15) *
                 (((((((v_x1 * dig_H6) >> 10) * (((v_x1 * dig_H3) >> 11) + 32768)) >> 10) + 2097152) *
                 dig_H2 + 8192) >> 14))
        v_x1 = v_x1 - (((((v_x1 >> 15) * (v_x1 >> 15)) >> 7) * dig_H1) >> 4)
        v_x1 = max(0, min(v_x1, 419430400))
        humidity = v_x1 >> 12
        return humidity / 1024.0
    ##############################################################################
    def __calculate_altitude( self , Tc , RH , P ):
        P0  = 1013.25     # 海面上の標準気圧[hPa]
        L   = 0.0065      # 温度減率[K/m]
        g   = 9.80665     # 重力加速度[m/s^2]
        R   = 8.314462618 # 気体定数[J/(mol·K)]
        M   = 0.0289644   # 空気のモル質量[kg/mol]
        Pu  = 226.32      # 標準大気における11kmの気圧[hPa]
        if RH is None or RH <= 0: # 湿度データが無い場合は実測温度を使う
            Tu = Tc + 273.15 # [K]
        else:                     # 湿度データがある場合は仮想温度を使う
            Tu = self.__virtual_temperature( Tc , RH , P )
        if P > Pu:                # 対流圏 (11km以下)
            h = (Tu / L) * (1 - (P / P0) ** ((R * L) / (g * M)))
        else:                     # 成層圏 (11km以上)
            h = 11000 + ( R * Tu ) / ( g * M ) * math.log( Pu / P )
        return h
    ##############################################################################
    def __virtual_temperature( self , Tc , RH , P ):
        Tk  = Tc + 273.15  # 気温をKに変換
        es  = 6.112 * math.exp( (17.67*Tc) / (Tc+243.5) ) # 飽和水蒸気圧(Tetensの式)es[hPa]
        e   = (RH / 100.0) * es                           # 実際の水蒸気圧e[hPa]
        r   = ((0.622*e) / (P-e)) / 1000                  # 混合比r[kg/kg]
        Tkv = Tk * ( 1 + 0.61 * r )                       # 仮想温度[K]
        return Tkv
    ##################################################################################
    # MPU6050
    ##################################################################################
    def __convert_mpu6050_dataFrame( self ):
        print("[Info] Start the __convert_mpu6050_dataFrame function.")
        # センサーレジスタデータを物理量データに変換
        self.__mergeDataFrame[
            [
                'mpu6050_ax'          ,
                'mpu6050_ay'          ,
                'mpu6050_az'          ,
                'mpu6050_gx'          ,
                'mpu6050_gy'          ,
                'mpu6050_gz'          ,
                'mpu6050_temperature'
            ]
        ] = self.__mergeDataFrame.apply(
            lambda row: pandas.Series(
                self.__convert_mpu6050_batch(
                    row['mpu6050_byte_00'] , row['mpu6050_byte_01'] , row['mpu6050_byte_02'] ,
                    row['mpu6050_byte_03'] , row['mpu6050_byte_04'] , row['mpu6050_byte_05'] ,
                    row['mpu6050_byte_06'] , row['mpu6050_byte_07'] , row['mpu6050_byte_08'] ,
                    row['mpu6050_byte_09'] , row['mpu6050_byte_10'] , row['mpu6050_byte_11'] ,
                    row['mpu6050_byte_12'] , row['mpu6050_byte_13']
                )
            ) , axis=1
        )
        self.__mergeDataFrame = self.__mergeDataFrame.drop(
            [
                'mpu6050_byte_00' , 'mpu6050_byte_01' , 'mpu6050_byte_02' ,
                'mpu6050_byte_03' , 'mpu6050_byte_04' , 'mpu6050_byte_05' ,
                'mpu6050_byte_06' , 'mpu6050_byte_07' , 'mpu6050_byte_08' ,
                'mpu6050_byte_09' , 'mpu6050_byte_10' , 'mpu6050_byte_11' ,
                'mpu6050_byte_12' , 'mpu6050_byte_13'
            ] ,
            axis=1
        )
    ##############################################################################
    def __convert_mpu6050_batch(
            self       ,
            byteData00 , byteData01 , byteData02 , byteData03 , byteData04 , byteData05 , byteData06 ,
            byteData07 , byteData08 , byteData09 , byteData10 , byteData11 , byteData12 , byteData13
    ):
        ax          = self.__convert_mpu6050( byteData00 , byteData01 )
        ay          = self.__convert_mpu6050( byteData02 , byteData03 )
        az          = self.__convert_mpu6050( byteData04 , byteData05 )
        temperature = self.__convert_mpu6050( byteData06 , byteData07 )
        gx          = self.__convert_mpu6050( byteData08 , byteData09 )
        gy          = self.__convert_mpu6050( byteData10 , byteData11 )
        gz          = self.__convert_mpu6050( byteData12 , byteData13 )
        return (
            ax/16384.0 , ay/16384.0 , az/16384.0 ,
            gx/131.0   , gy/131.0   , gz/131.0   ,
            temperature/340.0+36.53
        )
    ##############################################################################
    def __convert_mpu6050( self , msb , lsb ):
        value = struct.unpack('>h', bytes( [ msb , lsb ]))[0]
        return value
    ##################################################################################
    # ICM-20948
    ##################################################################################
    def __convert_icm20948_dataFrame( self ):
        print("[Info] Start the __convert_icm20948_dataFrame function.")
        # センサーレジスタデータを物理量データに変換
        self.__mergeDataFrame[
            [
                'icm-20948_ax'          ,
                'icm-20948_ay'          ,
                'icm-20948_az'          ,
                'icm-20948_gx'          ,
                'icm-20948_gy'          ,
                'icm-20948_gz'          ,
                'icm-20948_mx'          ,
                'icm-20948_my'          ,
                'icm-20948_mz'          ,
                'icm-20948_temperature' ,
                'icm-20948_heading_rad' ,
                'icm-20948_heading_deg' ,
                'icm-20948_crr_heading'
            ]
        ] = self.__mergeDataFrame.apply(
            lambda row: pandas.Series(
                self.__convert_icm20948_batch(
                    row['icm-20948_rawax'] , row['icm-20948_raway'] , row['icm-20948_rawaz'] ,
                    row['icm-20948_rawgx'] , row['icm-20948_rawgy'] , row['icm-20948_rawgz'] ,
                    row['icm-20948_rawmx'] , row['icm-20948_rawmy'] , row['icm-20948_rawmz'] ,
                    row['icm-20948_rawtemperature']
                )
            ) , axis=1
        )
        self.__mergeDataFrame = self.__mergeDataFrame.drop(
            [
                'icm-20948_rawax' , 'icm-20948_raway' , 'icm-20948_rawaz' ,
                'icm-20948_rawgx' , 'icm-20948_rawgy' , 'icm-20948_rawgz' ,
                'icm-20948_rawmx' , 'icm-20948_rawmy' , 'icm-20948_rawmz' ,
                'icm-20948_rawtemperature'
            ] ,
            axis=1
        )
    ##############################################################################
    def __convert_icm20948_batch( self , rawax , raway , rawaz , rawgx , rawgy , rawgz , rawmx , rawmy , rawmz , rawtemp ):
        temperature = rawtemp / 333.87 + 21
        if os.path.isfile( self.__calib_json ):
            f     = open( self.__calib_json , "r" )
            calib = json.load(f)
            f.close()
            offset           = numpy.array(calib["offset"])
            soft_iron_matrix = numpy.array(calib["soft_iron_matrix"])
            mx , my , mz = self.__apply_mag_calibration( rawmx , rawmy , rawmz , offset , soft_iron_matrix )
            accel_range = calib["accel_range"]
            gyro_range  = calib["gyro_range"]
            accel_scale = 32768.0
            if   accel_range == 250:
                accel_scale = 131.0
            elif accel_range == 500:
                accel_scale = 65.5
            elif accel_range == 1000:
                accel_scale = 32.8
            elif accel_range == 2000:
                accel_scale = 16.4
            gyro_scale = 131.0
            if   gyro_range == 250:
                gyro_scale = 131.0
            elif gyro_range == 500:
                gyro_scale = 65.5
            elif gyro_range == 1000:
                gyro_scale = 32.8
            elif gyro_range == 2000:
                gyro_scale = 16.4
            ax = rawax / accel_scale
            ay = raway / accel_scale
            az = rawaz / accel_scale
            gx = rawgx / gyro_scale
            gy = rawgy / gyro_scale
            gz = rawgz / gyro_scale
        else:
            mx = rawmx
            my = rawmy
            mz = rawmz
            ax = rawax / 32768.0
            ay = raway / 32768.0
            az = rawaz / 32768.0
            gx = rawgx / 131.0
            gy = rawgy / 131.0
            gz = rawgz / 131.0
        heading_rad = numpy.arctan2( my , mx )
        heading_deg = numpy.degrees( heading_rad )
        heading_deg = ( heading_deg + 360 ) % 360
        crr_heading = self.__calculate_tilt_compensated_heading( ax , ay , az , mx , my , mz )
        mx = mx * 0.15
        my = my * 0.15
        mz = mz * 0.15
        return ax , ay , az , gx , gy , gz , mx , my , mz , temperature , heading_rad , heading_deg , crr_heading
    ##############################################################################
    def __apply_mag_calibration( self , mx , my , mz , offset , soft_iron_matrix ):
        raw       = numpy.array([mx, my, mz])
        centered  = raw - offset
        corrected = soft_iron_matrix @ centered
        return corrected  # numpy array: [corrected_x, corrected_y, corrected_z]
    ##################################################################################
    def __normalize( self , v ):
        norm = numpy.linalg.norm(v)
        return v / norm if norm != 0 else v
    ##############################################################################
    def __calculate_tilt_compensated_heading( self , ax , ay , az , mx , my , mz ):
        acc   = self.__normalize(numpy.array([ ax , ay , az ]))
        pitch = numpy.arcsin( -acc[0] )
        roll  = numpy.arctan2( acc[1] , acc[2] )
        mx2   = mx * numpy.cos( pitch ) + mz * numpy.sin( pitch )
        my2   = (mx * numpy.sin( roll ) * numpy.sin( pitch ) +
                 my * numpy.cos( roll ) -
                 mz * numpy.sin( roll ) * numpy.cos( pitch ) )
        heading_rad = numpy.arctan2( my2 , mx2 )
        #heading_rad = numpy.arctan2( -my2 , mx2 )
        heading_deg = numpy.degrees( heading_rad )
        if heading_deg < 0:
            heading_deg += 360
        return heading_deg
    ##################################################################################
    # Power Monitor
    ##################################################################################
    def __convert_powermonitor_dataFrame( self ):
        print("[Info] Start the __convert_powermonitor_dataFrame function.")
        # センサーレジスタデータを物理量データに変換
        self.__mergeDataFrame[
            [
                'memory_usage_mb'         ,
                'memory_capacity_mb'      ,
                'free_memory_space_mb'    ,
                'cpu_temperature_c'       ,
                'disk_usage_gb'           ,
                'total_disk_capacity_gb'  ,
                'available_disk_space_gb'
            ]
        ] = self.__mergeDataFrame.apply(
            lambda row: pandas.Series(
                self.__convert_powermonitor_batch(
                    row['memory_usage']    , row['memory_capacity']     , row['free_memory_space']   ,
                    row['cpu_temperature'] ,
                    row['disk_usage']      , row['total_disk_capacity'] , row['available_disk_space']
                )
            ) , axis=1
        )
        self.__mergeDataFrame = self.__mergeDataFrame.drop(
            [
                'memory_usage'    , 'memory_capacity'     , 'free_memory_space'   ,
                'cpu_temperature' ,
                'disk_usage'      , 'total_disk_capacity' , 'available_disk_space'
            ] ,
            axis=1
        )
    ##################################################################################
    def __convert_powermonitor_batch(
            self ,
            memory_usage    , memory_capacity     , free_memory_space    ,
            cpu_temperature ,
            disk_usage      , total_disk_capacity , available_disk_space
    ):
        return (
            memory_usage/(1024*1024) , memory_capacity/(1024*1024)   , free_memory_space/(1024*1024)  ,
            cpu_temperature/1000     ,
            disk_usage/(1024**3)     , total_disk_capacity/(1024**3) , available_disk_space/(1024**3)
        )
    ##################################################################################
    def doI2CAnalyzerImpl(self):
        print("[Info] Start the doI2CAnalyzerImpl function.")
        self.__merge_csv()
        self.__mergeDataFrame['current_time'] = self.__mergeDataFrame['movie_end_unix_epoch_time'].apply(
            lambda epoch_time_ms :
            datetime.datetime.fromtimestamp(epoch_time_ms).strftime('%Y-%m-%d %H:%M:%S.') +
            f'{datetime.datetime.fromtimestamp(epoch_time_ms).microsecond // 1000:03d}'
        )
        if os.path.isfile( self.__input_dir + "/" + "powermonitor.csv" ):
            self.__convert_powermonitor_dataFrame()
        if os.path.isfile( self.__input_dir + "/" + "bme280.csv" ):
            self.__convert_bme280_dataFrame()
        if os.path.isfile( self.__input_dir + "/" + "mpu6050.csv" ):
            self.__convert_mpu6050_dataFrame()
        if os.path.isfile( self.__input_dir + "/" + "icm20948.csv" ):
            self.__convert_icm20948_dataFrame()
        if self.__excel_en:
            self.__output_to_excel( "analyse" , self.__input_dir + "/analyse.xlsx" , self.__mergeDataFrame )
        else:
            self.__mergeDataFrame.to_csv(  self.__input_dir + "/analyse.csv" , index=False )

########################################################################################
class GPSAnalyzerImpl:

    def __init__(self , csvFileName , animation_en ):
        self.__csvFileName  = csvFileName
        self.__animation_en = animation_en

    def __generate_map_html( self ):
        print("[Info] Start the __generate_map_html function.")
        dataFrame     = pandas.read_csv( self.__csvFileName )
        dataFrame     = dataFrame.reset_index()
        dataFrame["iso_8601_time"] = pandas.to_datetime(dataFrame["end_unix_epoch_time"], unit='s', utc=True).dt.strftime('%Y-%m-%dT%H:%M:%SZ')
        if self.__animation_en:
            features = []
            for _, row in dataFrame.iterrows():
                features.append(
                    { "type": "Feature", "geometry": { "type": "Point", "coordinates": [ row["longitude"], row["latitude"]], },
                      "properties": {
                          "time"      : row["iso_8601_time"]                      ,
                          "duration"  : 1000                                      ,
                          "popup"     : f"{row['latitude']} , {row['longitude']}" ,
                          "icon"      : "circle"                                  ,
                          "iconstyle" : {
                              "fillColor"   :"blue" ,
                              "fillOpacity" : 0.8   ,
                              "stroke"      :"true" ,
                              "radius"      : 6
                          },
                      }
                     })

            geojson    = { "type": "FeatureCollection", "features": features, }
            folium_map = folium.Map(
                location=[
                    dataFrame["ivk172_latitude"].iloc[0] ,
                    dataFrame["ivk172_latitude"].iloc[0]
                ],
                zoom_start=10
            )
            TimestampedGeoJson(
                geojson                  ,
                period         = "PT2S"  ,
                duration       = "PT0S"  ,
                add_last_point = False   ,
                auto_play      = True    ,
                loop           = True    ,
                max_speed      = 1       ,
            ).add_to(folium_map)
        else:
            folium_figure = folium.Figure(width=1500, height=700)
            folium_map    = folium.Map(
                location=[
                    dataFrame["ivk172_latitude"] .iloc[0] ,
                    dataFrame["ivk172_longitude"].iloc[0]
                ] ,
                zoom_start=4.5
            ).add_to( folium_figure )
            folium.PolyLine(
                dataFrame[["ivk172_latitude", "ivk172_longitude"]].values.tolist(),
                color="blue",
                weight=3,
                opacity=0.8
            ).add_to(folium_map)
            # for i in range( dataFrame.count()["latitude"] ):
            #     folium.Marker( location=[ dataFrame.loc[ i , "latitude" ] , dataFrame.loc[ i , "longitude" ] ] ).add_to( folium_map )
        folium_map.save( self.__csvFileName + ".html" )

    def __generate_map_kml( self ):
        print("[Info] Start the __generate_map_kml function.")
        dataFrame                        = pandas.read_csv( os.path.join(os.getcwd() , self.__csvFileName ) , header=0 )
        dataFrame                        = dataFrame.reset_index()
        tuple_B                          = [tuple(x) for x in dataFrame[['ivk172_longitude','ivk172_latitude','ivk172_altitude']].values]
        kml                              = simplekml.Kml(open=1)
        linestring                       = kml.newlinestring(name="A Sloped Line")
        linestring.coords                = tuple_B
        linestring.altitudemode          = simplekml.AltitudeMode.relativetoground
        linestring.extrude               = 0
        linestring.style.linestyle.width = 3
        linestring.style.linestyle.color = simplekml.Color.red
        kml.save( self.__csvFileName + ".kml" )

    def doGPSAnalyzerImpl( self ):
        print("[Info] Start the doGPSAnalyzerImpl function.")
        self.__generate_map_html()
        self.__generate_map_kml()

########################################################################################
class MovieAnalyzerImpl:

    def __init__( self , input_dir ):
        self.__input_dir = input_dir

    ############################################################################
    def __convert_h264_to_mp4( self , movieFileName ):
        print("[Info] Start the __convert_h264_to_mp4 function.")
        start_unix_epoch_time = time.time()
        if shutil.which("ffmpeg") is not None:
            if movieFileName is not None:
                print("[Info] Convert from H.264 to MP4.")
                print("[Info] ffmpeg -y -i " + movieFileName + " -c copy " + movieFileName + ".mp4" )
                subprocess.run(
                    "ffmpeg -y -i " + movieFileName + " -c copy " + movieFileName + ".mp4" ,
                    shell=True , capture_output=True , text=True
                )
            else:
                print("[Warn] Please set the video file name.")
        else:
            print("[Warn] Install it with the following command.")
            print("[Warn] apt install -y ffmpeg")
        end_unix_epoch_time = time.time()
        total_time = end_unix_epoch_time - start_unix_epoch_time
        print("[Info] The __convert_h264_to_mp4 function takes " + str(total_time) + " seconds to run.")
    ############################################################################
    def __separation_h264_to_jpeg( self , movieFileName ):
        print("[Info] Start the __separation_h264_to_jpeg function.")
        start_unix_epoch_time = time.time()
        if shutil.which("ffmpeg") is not None:
            print("[Info] ffmpeg -i " + movieFileName + " -qscale:v 2 tmp/frame_%08d.jpg")
            subprocess.run(
                "ffmpeg -i " + movieFileName + " -qscale:v 2 tmp/frame_%08d.jpg" ,
                shell          = True ,
                capture_output = True ,
                text           = True
            )
        else:
            print("[Warn] apt install -y ffmpeg")
        end_unix_epoch_time = time.time()
        total_time = end_unix_epoch_time - start_unix_epoch_time
        print("[Info] The __separation_h264_to_jpeg function takes " + str(total_time) + " seconds to run.")
    ############################################################################
    def __merge_jpeg_to_h264( self , movieFileName , framerate ):
        print("[Info] Start the __merge_jpeg_to_h264 function.")
        start_unix_epoch_time = time.time()
        if shutil.which("ffmpeg") is not None:
            print(
                "[Info] ffmpeg -framerate " + str(framerate) +
                " -i tmp/frame_opencv_%08d.jpg -c:v libx264 -f h264 -y " + movieFileName
            )
            subprocess.run(
                "ffmpeg -framerate " + str(framerate) +
                " -i tmp/frame_opencv_%08d.jpg -c:v libx264 -f h264 -y " + movieFileName ,
                shell          = True ,
                capture_output = True ,
                text           = True
            )
        else:
            print("[Warn] Install it with the following command.")
            print("[Warn] apt install -y ffmpeg")
        end_unix_epoch_time = time.time()
        total_time = end_unix_epoch_time - start_unix_epoch_time
        print("[Info] The __merge_jpeg_to_h264 function takes " + str(total_time) + " seconds to run.")

    #########################################################################
    def __add_sensor_frame( self , dataFrame , framerate ):
        print("[Info] Start the __add_sensor_frame function.")
        start_unix_epoch_time = time.time()
        imgFiles = sorted(glob.glob('tmp/frame_*.jpg'))
        frame_index = 0
        for imgFile in imgFiles:
            image    = cv2.imread(imgFile)
            text     =        "Date : " + str( dataFrame.iloc[frame_index]['current_time'] ) + "\n"
            text     = text + "Framerate : " + str( framerate ) + "\n"
            if os.path.isfile( self.__input_dir + "/" + "bme280.csv" ):
                text = text + "BME280 Altitude : "         + str( dataFrame.iloc[frame_index]['bme280_altitude']           ) + "\n"
                text = text + "BME280 Temperature : "      + str( dataFrame.iloc[frame_index]['bme280_temperature']        ) + "\n"
                text = text + "BME280 Pressure : "         + str( dataFrame.iloc[frame_index]['bme280_pressure']           ) + "\n"
                text = text + "BME280 Humidly : "          + str( dataFrame.iloc[frame_index]['bme280_humidity']           ) + "\n"
            if os.path.isfile( self.__input_dir + "/" + "mpu6050.csv" ):
                text = text + "MPU6050 AX : "              + str( dataFrame.iloc[frame_index]['mpu6050_ax']                ) + "\n"
                text = text + "MPU6050 AY : "              + str( dataFrame.iloc[frame_index]['mpu6050_ay']                ) + "\n"
                text = text + "MPU6050 AZ : "              + str( dataFrame.iloc[frame_index]['mpu6050_az']                ) + "\n"
                text = text + "MPU6050 GX : "              + str( dataFrame.iloc[frame_index]['mpu6050_gx']                ) + "\n"
                text = text + "MPU6050 GY : "              + str( dataFrame.iloc[frame_index]['mpu6050_gy']                ) + "\n"
                text = text + "MPU6050 GZ : "              + str( dataFrame.iloc[frame_index]['mpu6050_gz']                ) + "\n"
            if os.path.isfile( self.__input_dir + "/" + "icm20948.csv" ):
                text = text + "ICM20948 AX : "             + str( dataFrame.iloc[frame_index]['icm-20948_ax']              ) + "\n"
                text = text + "ICM20948 AY : "             + str( dataFrame.iloc[frame_index]['icm-20948_ay']              ) + "\n"
                text = text + "ICM20948 AZ : "             + str( dataFrame.iloc[frame_index]['icm-20948_az']              ) + "\n"
                text = text + "ICM20948 GX : "             + str( dataFrame.iloc[frame_index]['icm-20948_gx']              ) + "\n"
                text = text + "ICM20948 GY : "             + str( dataFrame.iloc[frame_index]['icm-20948_gy']              ) + "\n"
                text = text + "ICM20948 GZ : "             + str( dataFrame.iloc[frame_index]['icm-20948_gz']              ) + "\n"
                text = text + "ICM20948 MX : "             + str( dataFrame.iloc[frame_index]['icm-20948_mx']              ) + "\n"
                text = text + "ICM20948 MY : "             + str( dataFrame.iloc[frame_index]['icm-20948_my']              ) + "\n"
                text = text + "ICM20948 MZ : "             + str( dataFrame.iloc[frame_index]['icm-20948_mz']              ) + "\n"
                text = text + "ICM20948 heading rad : "    + str( dataFrame.iloc[frame_index]['icm-20948_heading_rad']     ) + "\n"
                text = text + "ICM20948 heading deg : "    + str( dataFrame.iloc[frame_index]['icm-20948_heading_deg']     ) + "\n"
            if os.path.isfile( self.__input_dir + "/" + "gps.csv" ):
                text = text + "GPS latitude : "            + str( dataFrame.iloc[frame_index]['ivk172_latitude']           ) + "\n"
                text = text + "GPS longitude : "           + str( dataFrame.iloc[frame_index]['ivk172_longitude']          ) + "\n"
                text = text + "GPS altitude : "            + str( dataFrame.iloc[frame_index]['ivk172_altitude']           ) + "\n"
                text = text + "GPS altitude_unit : "       + str( dataFrame.iloc[frame_index]['ivk172_altitude_units']     ) + "\n"
                text = text + "GPS num_sats : "            + str( dataFrame.iloc[frame_index]['ivk172_num_sats']           ) + "\n"
                text = text + "GPS datestam : "            + str( dataFrame.iloc[frame_index]['ivk172_datestam']           ) + "\n"
                text = text + "GPS timestamp: "            + str( dataFrame.iloc[frame_index]['ivk172_timestamp']          ) + "\n"
                text = text + "GPS spd over grnd : "       + str( dataFrame.iloc[frame_index]['ivk172_spd_over_grnd']      ) + "\n"
                text = text + "GPS true course : "         + str( dataFrame.iloc[frame_index]['ivk172_true_course']        ) + "\n"
                text = text + "GPS true track : "          + str( dataFrame.iloc[frame_index]['ivk172_true_track']         ) + "\n"
                text = text + "GPS spd over grnd kmph : "  + str( dataFrame.iloc[frame_index]['ivk172_spd_over_grnd_kmph'] ) + "\n"
                text = text + "GPS pdop : "                + str( dataFrame.iloc[frame_index]['ivk172_pdop']               ) + "\n"
                text = text + "GPS hdop : "                + str( dataFrame.iloc[frame_index]['ivk172_hdop']               ) + "\n"
                text = text + "GPS vdop : "                + str( dataFrame.iloc[frame_index]['ivk172_vdop']               ) + "\n"
                text = text + "GPS num sv in veiw : "      + str( dataFrame.iloc[frame_index]['ivk172_num_sv_in_view']     ) + "\n"
            if os.path.isfile( self.__input_dir + "/" + "powermonitor.csv" ):
                text = text + "voltage : "                    + str( dataFrame.iloc[frame_index]['voltage']                     ) + "\n"
                text = text + "throttled status : "           + str( dataFrame.iloc[frame_index]['throttled_status']            ) + "\n"
                text = text + "CPU utilization(%) : "         + str( dataFrame.iloc[frame_index]['cpu_utilization']             ) + "\n"
                text = text + "CPU Temperature(celsius) : "   + str( dataFrame.iloc[frame_index]['cpu_temperature_c']           ) + "\n"
                text = text + "Memory usage(MB) : "           + str( dataFrame.iloc[frame_index]['memory_usage_mb']             ) + "\n"
                text = text + "Memory Capacity(MB) : "        + str( dataFrame.iloc[frame_index]['memory_capacity_mb']          ) + "\n"
                text = text + "Free Memory Space(MB) : "      + str( dataFrame.iloc[frame_index]['free_memory_space_mb']        ) + "\n"
                text = text + "Memory usage Percentage(%) : " + str( dataFrame.iloc[frame_index]['memory_usage_percentage']     ) + "\n"
                text = text + "Disk usage(GB) : "             + str( dataFrame.iloc[frame_index]['disk_usage_gb']               ) + "\n"
                text = text + "Total Disk Capacity(GB) : "    + str( dataFrame.iloc[frame_index]['total_disk_capacity_gb']      ) + "\n"
                text = text + "Aveilable Disk Space(GB) : "   + str( dataFrame.iloc[frame_index]['available_disk_space_gb']     ) + "\n"
                text = text + "Disk utilization(%) : "        + str( dataFrame.iloc[frame_index]['disk_utilization_percentage'] ) + "\n"
            x , y       = 10 , 30
            #font        = cv2.FONT_HERSHEY_SIMPLEX
            font        = cv2.FONT_HERSHEY_PLAIN
            font_scale  = 2.25
            if os.path.isfile( self.__input_dir + "/" + "bme280.csv" ):
                font_scale = font_scale - 0.25                
            if os.path.isfile( self.__input_dir + "/" + "mpu6050.csv" ):
                font_scale = font_scale - 0.25
            if os.path.isfile( self.__input_dir + "/" + "icm20948.csv" ):
                font_scale = font_scale - 0.25
            if os.path.isfile( self.__input_dir + "/" + "gps.csv" ):
                font_scale = font_scale - 0.80
            if os.path.isfile( self.__input_dir + "/" + "powermonitor.csv" ):
                font_scale = font_scale - 0.25
            color       = ( 0 , 255 , 0 )
            thickness   = 1
            line_height = 30
            for i, line in enumerate(text.split('\n')):
                y_pos = y + i * line_height
                cv2.putText( image , line , (x, y_pos) , font , font_scale , color , thickness , cv2.LINE_AA )
            cv2.imwrite( str(re.sub(r"frame_", "frame_opencv_", imgFile )) , image)
            frame_index = frame_index + 1
        end_unix_epoch_time = time.time()
        total_time = end_unix_epoch_time - start_unix_epoch_time
        print("[Info] The __add_sensor_frame function takes " + str(total_time) + " seconds to run.")

    ############################################################################
    def __movie_gen( self , movieFileName , dataFrame ):
        print("[Info] Start the __movie_gen function.")
        shutil.rmtree( './tmp' , ignore_errors=True ) # remove tmp directory
        os.makedirs  ( './tmp' , exist_ok     =True ) # make tmp directory

        self.__separation_h264_to_jpeg( movieFileName )
        cap       = cv2.VideoCapture( movieFileName )
        framerate = cap.get( cv2.CAP_PROP_FPS )

        self.__add_sensor_frame   ( dataFrame , framerate )
        self.__merge_jpeg_to_h264 ( movieFileName + ".sensor.h264" , framerate )
        self.__convert_h264_to_mp4( movieFileName + ".sensor.h264" )
        shutil.rmtree( './tmp' , ignore_errors=True )

    #########################################################################
    def doMovieAnalyzerImpl( self , frame_sync_en , movieFileName , dataFrame ):
        if frame_sync_en :
            self.__movie_gen( movieFileName , dataFrame )
        else:
            self.__convert_h264_to_mp4( movieFileName )

########################################################################################
def main(argv):
    print("[Info] Start the main function.")
    sw = SensorWrapper( argv )
    sw.doSensorWrapper()

if __name__ == "__main__":
    sys.exit(main(sys.argv))
